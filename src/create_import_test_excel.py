from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def create_test_excel():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Import Test"

    # Normal values
    worksheet["A1"] = "Name"
    worksheet["B1"] = "Score"

    worksheet["A2"] = "Alice"
    worksheet["B2"] = 95

    worksheet["A3"] = "Bob"
    worksheet["B3"] = 87

    # Date
    worksheet["A4"] = "Date"
    worksheet["B4"] = date(2026, 9, 2)

    # Formula
    worksheet["A5"] = "Total"
    worksheet["B5"] = "=SUM(B2:B3)"

    # Meaningful formatting
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["B1"].font = Font(bold=True, size=14)

    worksheet["A1"].fill = PatternFill(
        fill_type="solid",
        fgColor="D9E1F2",
    )

    worksheet["B1"].fill = PatternFill(
        fill_type="solid",
        fgColor="D9E1F2",
    )

    worksheet["A1"].alignment = Alignment(horizontal="center")
    worksheet["B1"].alignment = Alignment(horizontal="center")

    # Intentionally empty/default cell.
    worksheet["H10"] = None

    file_path = "storage/import_test.xlsx"
    workbook.save(file_path)

    print(f"Test Excel file created: {file_path}")


if __name__ == "__main__":
    create_test_excel()