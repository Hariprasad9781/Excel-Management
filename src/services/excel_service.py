import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from models.excel_file import ExcelFile


# ============================================================
# File Helpers
# ============================================================


def get_excel_file_path(
    excel_file: ExcelFile,
) -> str:
    """
    Return the stored Excel file path.

    Raises:
        FileNotFoundError: If the stored Excel file does not exist.
    """

    if not os.path.exists(excel_file.file_path):
        raise FileNotFoundError(
            "Stored file not found"
        )

    return excel_file.file_path


def _load_workbook(
    excel_file: ExcelFile,
):
    """
    Load the Excel workbook from storage.
    """

    file_path = get_excel_file_path(
        excel_file
    )

    return load_workbook(file_path)


def _validate_sheet(
    workbook,
    sheet_name: str,
):
    """
    Validate that the requested sheet exists.
    """

    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found"
        )

    return workbook[sheet_name]


# ============================================================
# Sheet Operations
# ============================================================


def get_sheet_names(
    excel_file: ExcelFile,
) -> list[str]:
    """
    Return all sheet names in the workbook.
    """

    file_path = get_excel_file_path(
        excel_file
    )

    excel = pd.ExcelFile(file_path)

    return excel.sheet_names


def create_sheet(
    excel_file: ExcelFile,
    sheet_name: str,
) -> None:
    """
    Create a new worksheet.
    """

    file_path = get_excel_file_path(
        excel_file
    )

    if not sheet_name.strip():
        raise ValueError(
            "Sheet name cannot be empty"
        )

    workbook = load_workbook(file_path)

    if sheet_name in workbook.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' already exists"
        )

    workbook.create_sheet(
        sheet_name
    )

    workbook.save(file_path)


def rename_sheet(
    excel_file: ExcelFile,
    sheet_name: str,
    new_sheet_name: str,
) -> None:
    """
    Rename an existing worksheet.
    """

    file_path = get_excel_file_path(
        excel_file
    )

    workbook = load_workbook(file_path)

    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found"
        )

    if not new_sheet_name.strip():
        raise ValueError(
            "New sheet name cannot be empty"
        )

    if new_sheet_name in workbook.sheetnames:
        raise ValueError(
            f"Sheet '{new_sheet_name}' already exists"
        )

    worksheet = workbook[sheet_name]

    worksheet.title = new_sheet_name

    workbook.save(file_path)


def delete_sheet(
    excel_file: ExcelFile,
    sheet_name: str,
) -> None:
    """
    Delete a worksheet.

    The workbook must always contain at least one sheet.
    """

    file_path = get_excel_file_path(
        excel_file
    )

    workbook = load_workbook(file_path)

    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found"
        )

    if len(workbook.sheetnames) == 1:
        raise ValueError(
            "Cannot delete the only sheet in the workbook"
        )

    worksheet = workbook[sheet_name]

    workbook.remove(worksheet)

    workbook.save(file_path)


def get_sheet_preview(
    excel_file: ExcelFile,
    sheet_name: str,
    rows: int = 10,
) -> dict:
    """
    Return a preview of the requested worksheet.
    """

    file_path = get_excel_file_path(
        excel_file
    )

    excel = pd.ExcelFile(file_path)

    if sheet_name not in excel.sheet_names:
        raise ValueError(
            f"Sheet '{sheet_name}' not found"
        )

    df = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        nrows=rows,
    )

    df = df.fillna("")

    return {
        "sheet_name": sheet_name,
        "columns": df.columns.tolist(),
        "rows": df.to_dict(
            orient="records"
        ),
        "row_count": len(df),
    }


# ============================================================
# Cell Operations
# ============================================================


def update_cell(
    excel_file: ExcelFile,
    sheet_name: str,
    cell: str,
    value: str | int | float | bool | None,
) -> None:
    """
    Update the value of a single Excel cell.
    """

    file_path = get_excel_file_path(
        excel_file
    )

    workbook = load_workbook(file_path)

    worksheet = _validate_sheet(
        workbook,
        sheet_name,
    )

    if not cell.strip():
        raise ValueError(
            "Cell reference cannot be empty"
        )

    try:
        worksheet[cell] = value
    except ValueError as exc:
        raise ValueError(
            f"Invalid cell reference: {cell}"
        ) from exc

    workbook.save(file_path)


# ============================================================
# Row Operations
# ============================================================


def add_row(
    excel_file: ExcelFile,
    sheet_name: str,
    row_data: list[
        str | int | float | bool | None
    ],
) -> int:
    """
    Add a row to the end of the worksheet.

    Returns:
        The Excel row number that was added.
    """

    file_path = get_excel_file_path(
        excel_file
    )

    workbook = load_workbook(file_path)

    worksheet = _validate_sheet(
        workbook,
        sheet_name,
    )

    if not row_data:
        raise ValueError(
            "Row data cannot be empty"
        )

    worksheet.append(row_data)

    row_number = worksheet.max_row

    workbook.save(file_path)

    return row_number


def update_row(
    excel_file: ExcelFile,
    sheet_name: str,
    row_number: int,
    row_data: list[
        str | int | float | bool | None
    ],
) -> None:
    """
    Update an existing Excel row.
    """

    file_path = get_excel_file_path(
        excel_file
    )

    workbook = load_workbook(file_path)

    worksheet = _validate_sheet(
        workbook,
        sheet_name,
    )

    if row_number < 1:
        raise ValueError(
            "Row number must be greater than 0"
        )

    if row_number > worksheet.max_row:
        raise ValueError(
            f"Row {row_number} not found"
        )

    if not row_data:
        raise ValueError(
            "Row data cannot be empty"
        )

    for column_number, value in enumerate(
        row_data,
        start=1,
    ):
        worksheet.cell(
            row=row_number,
            column=column_number,
            value=value,
        )

    workbook.save(file_path)


def delete_row(
    excel_file: ExcelFile,
    sheet_name: str,
    row_number: int,
) -> None:
    """
    Delete an existing Excel row.
    """

    file_path = get_excel_file_path(
        excel_file
    )

    workbook = load_workbook(file_path)

    worksheet = _validate_sheet(
        workbook,
        sheet_name,
    )

    if row_number < 1:
        raise ValueError(
            "Row number must be greater than 0"
        )

    if row_number > worksheet.max_row:
        raise ValueError(
            f"Row {row_number} not found"
        )

    worksheet.delete_rows(
        row_number,
        1,
    )

    workbook.save(file_path)


# ============================================================
# Column Operations
# ============================================================


def add_column(
    excel_file: ExcelFile,
    sheet_name: str,
    column_number: int,
) -> None:
    """
    Insert a new blank column.

    The blank cells are explicitly created so that the
    inserted column remains part of the worksheet dimensions.
    """

    file_path = get_excel_file_path(
        excel_file
    )

    workbook = load_workbook(file_path)

    worksheet = _validate_sheet(
        workbook,
        sheet_name,
    )

    if column_number < 1:
        raise ValueError(
            "Column number must be greater than 0"
        )

    max_column = worksheet.max_column

    if column_number > max_column + 1:
        raise ValueError(
            f"Column {column_number} is out of range. "
            f"Next available column is {max_column + 1}."
        )

    worksheet.insert_cols(
        column_number,
        1,
    )

    row_count = max(
        worksheet.max_row,
        1,
    )

    for row_number in range(
        1,
        row_count + 1,
    ):
        worksheet.cell(
            row=row_number,
            column=column_number,
            value="",
        )

    workbook.save(file_path)


def update_column(
    excel_file: ExcelFile,
    sheet_name: str,
    column_number: int,
    column_name: str,
) -> None:
    """
    Update the header/name of a column.
    """

    file_path = get_excel_file_path(
        excel_file
    )

    workbook = load_workbook(file_path)

    worksheet = _validate_sheet(
        workbook,
        sheet_name,
    )

    if column_number < 1:
        raise ValueError(
            "Column number must be greater than 0"
        )

    if column_number > worksheet.max_column:
        raise ValueError(
            f"Column {column_number} not found"
        )

    if not column_name.strip():
        raise ValueError(
            "Column name cannot be empty"
        )

    worksheet.cell(
        row=1,
        column=column_number,
        value=column_name,
    )

    workbook.save(file_path)


def delete_column(
    excel_file: ExcelFile,
    sheet_name: str,
    column_number: int,
) -> None:
    """
    Delete an existing column.
    """

    file_path = get_excel_file_path(
        excel_file
    )

    workbook = load_workbook(file_path)

    worksheet = _validate_sheet(
        workbook,
        sheet_name,
    )

    if column_number < 1:
        raise ValueError(
            "Column number must be greater than 0"
        )

    if column_number > worksheet.max_column:
        raise ValueError(
            f"Column {column_number} not found"
        )

    worksheet.delete_cols(
        column_number,
        1,
    )

    workbook.save(file_path)


# ============================================================
# Excel Search
# ============================================================


def search_excel(
    excel_file: ExcelFile,
    sheet_name: str,
    search_term: str,
) -> list[dict]:
    """
    Search the complete worksheet for a term.

    Returns:
        A list containing matching cells and their
        row/column information.

    Note:
        row_number is the actual Excel row number.
        For example, if the header is on row 1 and
        'Hari' is in A2, row_number will be 2.
    """

    file_path = get_excel_file_path(
        excel_file
    )

    workbook = load_workbook(
        file_path,
        data_only=False,
    )

    worksheet = _validate_sheet(
        workbook,
        sheet_name,
    )

    if not search_term.strip():
        raise ValueError(
            "Search term cannot be empty"
        )

    results = []

    search_value = search_term.lower()

    for row in worksheet.iter_rows():

        for cell in row:

            if cell.value is None:
                continue

            cell_value = str(
                cell.value
            )

            if search_value in cell_value.lower():
                results.append(
                    {
                        "row_number": cell.row,
                        "column_number": cell.column,
                        "cell": cell.coordinate,
                        "value": cell.value,
                    }
                )

    return results


# ============================================================
# Excel Formatting
# ============================================================


def format_excel_range(
    excel_file: ExcelFile,
    sheet_name: str,
    cell_range: str,
    bold: bool | None = None,
    italic: bool | None = None,
    underline: bool | None = None,
    font_size: float | None = None,
    font_color: str | None = None,
    fill_color: str | None = None,
    horizontal_alignment: str | None = None,
    vertical_alignment: str | None = None,
    number_format: str | None = None,
) -> None:
    """
    Apply formatting to a single cell or a cell range.

    Examples:
        A1
        B2
        A1:D5
    """

    file_path = get_excel_file_path(
        excel_file
    )

    workbook = load_workbook(file_path)

    worksheet = _validate_sheet(
        workbook,
        sheet_name,
    )

    if not cell_range.strip():
        raise ValueError(
            "Cell range cannot be empty"
        )

    # --------------------------------------------------------
    # Validate the cell range
    # --------------------------------------------------------

    try:
        selected_range = worksheet[cell_range]
    except (ValueError, KeyError) as exc:
        raise ValueError(
            f"Invalid cell range: {cell_range}"
        ) from exc

    # --------------------------------------------------------
    # Convert single cell / range into a list of cells
    # --------------------------------------------------------

    if hasattr(selected_range, "coordinate"):
        # Single cell, e.g. A1
        cells = [selected_range]

    elif isinstance(selected_range, tuple):

        if not selected_range:
            raise ValueError(
                f"Invalid cell range: {cell_range}"
            )

        # Range, e.g. A1:D5
        if isinstance(
            selected_range[0],
            tuple,
        ):
            cells = [
                cell
                for row in selected_range
                for cell in row
            ]
        else:
            cells = list(selected_range)

    else:
        raise ValueError(
            f"Invalid cell range: {cell_range}"
        )

    # --------------------------------------------------------
    # Validate alignment values
    # --------------------------------------------------------

    valid_horizontal_alignments = {
        "left",
        "center",
        "right",
        "fill",
        "justify",
        "centerContinuous",
        "distributed",
    }

    valid_vertical_alignments = {
        "top",
        "center",
        "bottom",
        "justify",
        "distributed",
    }

    if (
        horizontal_alignment is not None
        and horizontal_alignment
        not in valid_horizontal_alignments
    ):
        raise ValueError(
            "Invalid horizontal alignment. "
            "Allowed values: "
            + ", ".join(
                sorted(valid_horizontal_alignments)
            )
        )

    if (
        vertical_alignment is not None
        and vertical_alignment
        not in valid_vertical_alignments
    ):
        raise ValueError(
            "Invalid vertical alignment. "
            "Allowed values: "
            + ", ".join(
                sorted(valid_vertical_alignments)
            )
        )

    # --------------------------------------------------------
    # Validate font size
    # --------------------------------------------------------

    if font_size is not None and font_size <= 0:
        raise ValueError(
            "Font size must be greater than 0"
        )

    # --------------------------------------------------------
    # Normalize colors
    # --------------------------------------------------------

    def normalize_color(
        color: str,
    ) -> str:
        """
        Convert a 6-digit or 8-digit hex color
        into an 8-digit ARGB value.
        """

        normalized = color.strip().replace(
            "#",
            "",
        ).upper()

        if len(normalized) == 6:
            normalized = "FF" + normalized

        if len(normalized) != 8:
            raise ValueError(
                "Color must be a 6-digit or 8-digit "
                "hexadecimal value"
            )

        valid_hex = all(
            character in "0123456789ABCDEF"
            for character in normalized
        )

        if not valid_hex:
            raise ValueError(
                "Color must contain only hexadecimal "
                "characters"
            )

        return normalized

    normalized_font_color = None

    if font_color is not None:
        normalized_font_color = normalize_color(
            font_color
        )

    normalized_fill_color = None

    if fill_color is not None:
        normalized_fill_color = normalize_color(
            fill_color
        )

    # --------------------------------------------------------
    # Apply formatting
    # --------------------------------------------------------

    for cell in cells:

        # ----------------------------------------------------
        # Font formatting
        # ----------------------------------------------------

        if any(
            value is not None
            for value in (
                bold,
                italic,
                underline,
                font_size,
                normalized_font_color,
            )
        ):

            current_font = cell.font

            cell.font = Font(
                name=current_font.name,
                sz=(
                    font_size
                    if font_size is not None
                    else current_font.sz
                ),
                bold=(
                    bold
                    if bold is not None
                    else current_font.bold
                ),
                italic=(
                    italic
                    if italic is not None
                    else current_font.italic
                ),
                underline=(
                    underline
                    if underline is not None
                    else current_font.underline
                ),
                strike=current_font.strike,
                color=(
                    normalized_font_color
                    if normalized_font_color is not None
                    else current_font.color
                ),
                vertAlign=current_font.vertAlign,
                charset=current_font.charset,
                family=current_font.family,
                scheme=current_font.scheme,
                outline=current_font.outline,
                shadow=current_font.shadow,
                condense=current_font.condense,
                extend=current_font.extend,
            )

        # ----------------------------------------------------
        # Background fill
        # ----------------------------------------------------

        if normalized_fill_color is not None:
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=normalized_fill_color,
            )

        # ----------------------------------------------------
        # Alignment
        # ----------------------------------------------------

        if (
            horizontal_alignment is not None
            or vertical_alignment is not None
        ):

            current_alignment = cell.alignment

            cell.alignment = Alignment(
                horizontal=(
                    horizontal_alignment
                    if horizontal_alignment is not None
                    else current_alignment.horizontal
                ),
                vertical=(
                    vertical_alignment
                    if vertical_alignment is not None
                    else current_alignment.vertical
                ),
                textRotation=current_alignment.textRotation,
                wrapText=current_alignment.wrapText,
                shrinkToFit=current_alignment.shrinkToFit,
                indent=current_alignment.indent,
            )

        # ----------------------------------------------------
        # Number format
        # ----------------------------------------------------

        if number_format is not None:
            if not number_format.strip():
                raise ValueError(
                    "Number format cannot be empty"
                )

            cell.number_format = number_format

    # --------------------------------------------------------
    # Save workbook
    # --------------------------------------------------------

    workbook.save(file_path)