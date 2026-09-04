from __future__ import annotations

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from models.excel_file import ExcelFile
from models.workbook import Workbook
from models.workbook_version import WorkbookVersion
from services.excel_service import get_excel_file_path


def _serialize_value(value):
    """
    Convert Excel values into JSON-safe values.
    """

    if value is None:
        return None

    if hasattr(value, "isoformat"):
        return value.isoformat()

    if isinstance(value, (str, int, float, bool)):
        return value

    return str(value)


def _serialize_style(cell) -> dict:
    """
    Capture the basic formatting of a cell
    using JSON-safe values.
    """

    def serialize_color(color):
        if color is None:
            return None

        return {
            "type": str(color.type) if color.type is not None else None,
            "rgb": str(color.rgb) if color.rgb is not None else None,
            "indexed": str(color.indexed) if color.indexed is not None else None,
            "theme": str(color.theme) if color.theme is not None else None,
            "tint": str(color.tint) if color.tint is not None else None,
            "auto": str(color.auto) if color.auto is not None else None,
        }

    return {
        "font": {
            "name": cell.font.name,
            "size": cell.font.sz,
            "bold": cell.font.bold,
            "italic": cell.font.italic,
            "underline": cell.font.underline,
            "strike": cell.font.strike,
        },
        "fill": {
            "fill_type": cell.fill.fill_type,
            "fgColor": serialize_color(cell.fill.fgColor),
            "bgColor": serialize_color(cell.fill.bgColor),
        },
        "alignment": {
            "horizontal": cell.alignment.horizontal,
            "vertical": cell.alignment.vertical,
            "wrap_text": cell.alignment.wrap_text,
        },
        "number_format": cell.number_format,
    }


def _build_snapshot(excel_file: ExcelFile) -> dict:
    """
    Build a complete snapshot from the current physical Excel file.
    """

    file_path = get_excel_file_path(excel_file)

    workbook = load_workbook(
        file_path,
        data_only=False,
    )

    worksheets_snapshot = []

    for position, worksheet in enumerate(
        workbook.worksheets,
        start=1,
    ):
        cells_snapshot = []

        for row in worksheet.iter_rows():
            for cell in row:

                # Ignore completely empty cells without formatting.
                if (
                    cell.value is None
                    and not cell.has_style
                ):
                    continue

                value = cell.value

                formula = None

                if (
                    isinstance(value, str)
                    and value.startswith("=")
                ):
                    formula = value

                cells_snapshot.append(
                    {
                        "row_index": cell.row,
                        "column_index": cell.column,
                        "value": _serialize_value(value),
                        "data_type": cell.data_type,
                        "formula": formula,
                        "style": _serialize_style(cell),
                    }
                )

        worksheets_snapshot.append(
            {
                "name": worksheet.title,
                "position": position,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "cells": cells_snapshot,
            }
        )

    return {
        "workbook": {
            "id": excel_file.workbook.id,
            "name": excel_file.workbook.name,
            "original_filename": excel_file.original_filename,
        },
        "worksheets": worksheets_snapshot,
    }


def create_file_version(
    db: Session,
    excel_file: ExcelFile,
    created_by: int,
    change_summary: str | None = None,
) -> WorkbookVersion:
    """
    Create a new version from the current physical Excel file.

    The Excel file must already contain the latest changes.
    """

    if excel_file.workbook_id is None:
        raise ValueError(
            "Excel file is not linked to a workbook."
        )

    workbook = (
        db.query(Workbook)
        .filter(
            Workbook.id == excel_file.workbook_id,
            Workbook.is_deleted.is_(False),
        )
        .first()
    )

    if workbook is None:
        raise ValueError(
            f"Workbook {excel_file.workbook_id} not found."
        )

    # Build snapshot from the actual .xlsx file.
    snapshot = _build_snapshot(excel_file)

    next_version = workbook.version + 1

    version = WorkbookVersion(
        workbook_id=workbook.id,
        version_number=next_version,
        created_by=created_by,
        snapshot_data=snapshot,
        change_summary=change_summary,
    )

    db.add(version)

    workbook.version = next_version

    return version

def restore_file_version(
    db: Session,
    excel_file: ExcelFile,
    version_number: int,
) -> WorkbookVersion:
    """
    Restore a workbook version into the physical Excel file.

    The selected version snapshot is used to rebuild the .xlsx file.
    """

    if excel_file.workbook_id is None:
        raise ValueError(
            "Excel file is not linked to a workbook."
        )

    version = (
        db.query(WorkbookVersion)
        .filter(
            WorkbookVersion.workbook_id == excel_file.workbook_id,
            WorkbookVersion.version_number == version_number,
        )
        .first()
    )

    if version is None:
        raise ValueError(
            f"Version {version_number} not found."
        )

    snapshot_data = version.snapshot_data

    if not snapshot_data:
        raise ValueError(
            f"Version {version_number} does not contain snapshot data."
        )

    worksheets_snapshot = snapshot_data.get(
        "worksheets",
        [],
    )

    if not worksheets_snapshot:
        raise ValueError(
            f"Version {version_number} contains no worksheets."
        )

    # Create a new Excel workbook.
    from openpyxl import Workbook as OpenPyXLWorkbook

    excel_workbook = OpenPyXLWorkbook()

    # Remove the default worksheet.
    default_sheet = excel_workbook.active
    excel_workbook.remove(default_sheet)

    for worksheet_data in worksheets_snapshot:
        worksheet = excel_workbook.create_sheet(
            title=worksheet_data["name"],
        )

        cells_snapshot = worksheet_data.get(
            "cells",
            [],
        )

        for cell_data in cells_snapshot:
            cell = worksheet.cell(
                row=cell_data["row_index"],
                column=cell_data["column_index"],
            )

            cell.value = cell_data.get("value")

            style = cell_data.get("style")

            if not style:
                continue

            font_data = style.get("font", {})

            cell.font = cell.font.copy(
                name=font_data.get("name"),
                size=font_data.get("size"),
                bold=font_data.get("bold"),
                italic=font_data.get("italic"),
                underline=font_data.get("underline"),
                strike=font_data.get("strike"),
            )

            alignment_data = style.get(
                "alignment",
                {},
            )

            cell.alignment = cell.alignment.copy(
                horizontal=alignment_data.get("horizontal"),
                vertical=alignment_data.get("vertical"),
                wrap_text=alignment_data.get("wrap_text"),
            )

            number_format = style.get("number_format")

            if number_format:
                cell.number_format = number_format

        # Preserve worksheet dimensions where possible.
        max_row = worksheet_data.get("max_row")
        max_column = worksheet_data.get("max_column")

        if max_row:
            worksheet.sheet_view.showGridLines = True

    file_path = get_excel_file_path(excel_file)

    excel_workbook.save(file_path)
    excel_workbook.close()

    return version