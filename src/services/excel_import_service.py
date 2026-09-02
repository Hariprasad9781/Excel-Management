from __future__ import annotations

import os
from datetime import date, datetime
from decimal import Decimal

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from models.cell import Cell
from models.workbook import Workbook
from models.worksheet import Worksheet


SUPPORTED_EXTENSION = ".xlsx"


def _serialize_cell_value(value):
    """
    Convert Excel cell values into JSON-compatible values.
    """

    if value is None:
        return None

    if isinstance(value, datetime):
        return value.isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, Decimal):
        return float(value)

    if isinstance(value, (str, int, float, bool)):
        return value

    return str(value)


def _get_cell_data_type(cell) -> str | None:
    """
    Determine a simple application-level data type.
    """

    if cell.value is None:
        return None

    if cell.data_type == "f":
        return "formula"

    if cell.is_date:
        return "date"

    if isinstance(cell.value, bool):
        return "boolean"

    if isinstance(cell.value, (int, float)):
        return "number"

    if isinstance(cell.value, str):
        return "string"

    return "string"


def _normalize_color(color) -> str | None:
    """
    Convert an openpyxl color into an ARGB hex string when possible.
    """

    if color is None:
        return None

    color_type = color.type

    if color_type == "rgb" and color.rgb:
        return color.rgb

    return None


def _extract_cell_style(cell) -> dict:
    """
    Extract the formatting currently supported by the application.
    """

    style = {}

    if cell.font:
        if cell.font.bold is not None:
            style["bold"] = cell.font.bold

        if cell.font.italic is not None:
            style["italic"] = cell.font.italic

        if cell.font.underline:
            style["underline"] = cell.font.underline

        if cell.font.sz is not None:
            style["font_size"] = cell.font.sz

        font_color = _normalize_color(cell.font.color)

        if font_color:
            style["font_color"] = font_color

    if cell.fill:
        fill_color = _normalize_color(cell.fill.fgColor)

        if fill_color and fill_color not in {"00000000", "000000"}:
            style["fill_color"] = fill_color

    if cell.alignment:
        if cell.alignment.horizontal:
            style["horizontal_alignment"] = cell.alignment.horizontal

        if cell.alignment.vertical:
            style["vertical_alignment"] = cell.alignment.vertical

    if cell.number_format:
        if cell.number_format != "General":
            style["number_format"] = cell.number_format

    return style


def _cell_has_content_or_style(cell) -> bool:
    # Always keep cells that contain a value or formula.
    if cell.value is not None:
        return True

    # Ignore Excel's default formatting.
    style = _extract_cell_style(cell)

    meaningful_style = {
        key: value
        for key, value in style.items()
        if value not in (None, False, 11.0, "General")
    }

    return bool(meaningful_style)


def import_excel_to_database(
    db: Session,
    file_path: str,
    owner_id: int,
    original_filename: str | None = None,
) -> Workbook:
    """
    Import an .xlsx workbook into PostgreSQL.

    Excel structure:

        Workbook
            └── Worksheets
                    └── Cells

    The entire operation runs inside one database transaction.
    If anything fails, no partial workbook is committed.
    """

    if not file_path:
        raise ValueError("File path is required")

    if not os.path.exists(file_path):
        raise FileNotFoundError(
            f"Excel file not found: {file_path}"
        )

    extension = os.path.splitext(file_path)[1].lower()

    if extension != SUPPORTED_EXTENSION:
        raise ValueError(
            "Only .xlsx files are supported for database import"
        )

    if original_filename:
        workbook_name = os.path.splitext(
            original_filename
        )[0]
    else:
        workbook_name = os.path.splitext(
            os.path.basename(file_path)
        )[0]

    workbook = None

    try:
        # ----------------------------------------------------
        # Load Excel workbook
        # ----------------------------------------------------

        excel_workbook = load_workbook(
            filename=file_path,
            data_only=False,
        )

        # ----------------------------------------------------
        # Create database workbook
        # ----------------------------------------------------

        workbook = Workbook(
            owner_id=owner_id,
            name=workbook_name,
            original_filename=original_filename,
            version=1,
            is_deleted=False,
        )

        db.add(workbook)
        db.flush()

        # ----------------------------------------------------
        # Import worksheets
        # ----------------------------------------------------

        for position, excel_sheet in enumerate(
            excel_workbook.worksheets
        ):
            worksheet = Worksheet(
                workbook_id=workbook.id,
                name=excel_sheet.title,
                position=position,
                max_row=excel_sheet.max_row or 0,
                max_column=excel_sheet.max_column or 0,
                is_deleted=False,
            )

            db.add(worksheet)
            db.flush()

            # ------------------------------------------------
            # Import meaningful cells
            # ------------------------------------------------

            cells_to_insert = []

            for row in excel_sheet.iter_rows():
                for excel_cell in row:

                    if not _cell_has_content_or_style(
                        excel_cell
                    ):
                        continue

                    value = excel_cell.value

                    formula = None

                    if excel_cell.data_type == "f":
                        formula = value

                    cell = Cell(
                        worksheet_id=worksheet.id,
                        row_index=excel_cell.row,
                        column_index=excel_cell.column,
                        value=_serialize_cell_value(
                            value
                        ),
                        data_type=_get_cell_data_type(
                            excel_cell
                        ),
                        formula=formula,
                        style=_extract_cell_style(
                            excel_cell
                        ) or None,
                    )

                    cells_to_insert.append(cell)

            if cells_to_insert:
                db.add_all(cells_to_insert)

        # ----------------------------------------------------
        # Commit entire import
        # ----------------------------------------------------

        db.commit()

        db.refresh(workbook)

        return workbook

    except Exception:
        db.rollback()
        raise

    finally:
        try:
            excel_workbook.close()
        except (
            UnboundLocalError,
            AttributeError,
        ):
            pass