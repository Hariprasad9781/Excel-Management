import os
import tempfile
from copy import copy

from PySide6.QtCore import Qt
from PySide6.QtWidgets import QApplication
from PySide6.QtGui import QColor, QFont
from PySide6.QtWidgets import (
    QColorDialog,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QInputDialog,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QPushButton,
    QSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)
from openpyxl import load_workbook


class ExcelEditorWindow(QWidget):
    """
    Excel editor window.

    The backend remains the source of truth for every structural
    operation. The local workbook/table is updated only after the
    backend operation succeeds.

    Supported operations:
        - Edit cells
        - Save cell changes
        - Add row
        - Delete row
        - Add column
        - Rename column
        - Delete column
        - Create sheet
        - Rename sheet
        - Delete sheet
        - Search workbook
        - Font and fill formatting
        - Alignment formatting
        - Number formatting
        - Table-style formatting preset
        - Format Cells dialog
        - Refresh workbook
    """

    def __init__(
        self,
        file_id: int,
        filename: str,
        api_client,
    ):
        super().__init__()

        self.file_id = file_id
        self.filename = filename
        self.api_client = api_client

        self.workbook = None
        self.temp_file_path = None

        # Unsaved cell changes only.
        # {(sheet_name, "A1"): value}
        self.pending_changes = {}

        # Prevents cellChanged from treating programmatic table
        # updates as user edits.
        self._loading_table = False

        # Formatting state.
        # The backend accepts underline styles as strings.
        self.underline_style = None

        # Search results returned by the backend.
        self.search_results = []

        self.setWindowTitle(
            f"Excel Editor - {self.filename}"
        )
        self.setMinimumSize(1200, 700)

        self.setup_ui()
        self.load_excel_file()

        # Buttons
        self.add_row_button.clicked.connect(self.add_row)
        self.delete_row_button.clicked.connect(self.delete_row)
        self.add_column_button.clicked.connect(self.add_column)
        self.rename_column_button.clicked.connect(self.rename_column)
        self.delete_column_button.clicked.connect(self.delete_column)
        self.save_button.clicked.connect(self.save_changes)

        # Sheet management
        self.create_sheet_button.clicked.connect(self.create_sheet)
        self.rename_sheet_button.clicked.connect(self.rename_sheet)
        self.delete_sheet_button.clicked.connect(self.delete_sheet)
        self.version_history_button.clicked.connect(self.show_version_history)

        # Formatting controls
        self.bold_button.clicked.connect(self.toggle_bold)
        self.italic_button.clicked.connect(self.toggle_italic)
        self.underline_button.clicked.connect(self.toggle_underline)
        self.font_size_spin.valueChanged.connect(self.change_font_size)
        self.font_color_button.clicked.connect(self.choose_font_color)
        self.fill_color_button.clicked.connect(self.choose_fill_color)
        self.left_align_button.clicked.connect(
            lambda: self.apply_alignment("left")
        )
        self.center_align_button.clicked.connect(
            lambda: self.apply_alignment("center")
        )
        self.right_align_button.clicked.connect(
            lambda: self.apply_alignment("right")
        )
        self.top_align_button.clicked.connect(
            lambda: self.apply_vertical_alignment("top")
        )
        self.middle_align_button.clicked.connect(
            lambda: self.apply_vertical_alignment("center")
        )
        self.bottom_align_button.clicked.connect(
            lambda: self.apply_vertical_alignment("bottom")
        )
        self.number_format_combo.currentTextChanged.connect(
            self.change_number_format
        )
        self.table_format_button.clicked.connect(self.apply_table_format)
        self.clear_format_button.clicked.connect(self.clear_format)
        self.format_cells_button.clicked.connect(self.open_format_cells_dialog)

    # =========================================================
    # UI
    # =========================================================

    def setup_ui(self):
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(15)

        header_layout = QHBoxLayout()

        title = QLabel(
            f"Excel Editor - {self.filename}"
        )
        title.setStyleSheet(
            """
            QLabel {
                font-size: 24px;
                font-weight: bold;
            }
            """
        )

        header_layout.addWidget(title)
        header_layout.addStretch()

        header_layout.addWidget(QLabel("Sheet:"))

        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(180)
        self.sheet_combo.currentTextChanged.connect(
            self.load_sheet
        )
        header_layout.addWidget(self.sheet_combo)

        self.create_sheet_button = QPushButton("New Sheet")
        self.rename_sheet_button = QPushButton("Rename Sheet")
        self.delete_sheet_button = QPushButton("Delete Sheet")
        self.version_history_button = QPushButton("Version History")
        self.refresh_button = QPushButton("Refresh")

        for button in (
            self.create_sheet_button,
            self.rename_sheet_button,
            self.delete_sheet_button,
            self.version_history_button,
            self.refresh_button,
        ):
            button.setMinimumHeight(38)

        header_layout.addWidget(self.create_sheet_button)
        header_layout.addWidget(self.rename_sheet_button)
        header_layout.addWidget(self.delete_sheet_button)
        header_layout.addWidget(self.version_history_button)
        header_layout.addWidget(self.refresh_button)
        self.refresh_button.clicked.connect(
            self.load_excel_file
        )

        main_layout.addLayout(header_layout)

        # =====================================================
        # Search
        # =====================================================

        search_layout = QHBoxLayout()

        search_layout.addWidget(QLabel("Search:"))

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText(
            "Search in the selected sheet..."
        )
        self.search_input.setMinimumHeight(36)

        self.search_button = QPushButton("Search")
        self.clear_search_button = QPushButton("Clear")

        self.search_button.setMinimumHeight(36)
        self.clear_search_button.setMinimumHeight(36)

        self.search_result_label = QLabel(
            "No search performed."
        )

        search_layout.addWidget(self.search_input, 1)
        search_layout.addWidget(self.search_button)
        search_layout.addWidget(self.clear_search_button)
        search_layout.addWidget(self.search_result_label)

        main_layout.addLayout(search_layout)

        # =====================================================
        # Search Results
        # =====================================================

        self.search_results_list = QListWidget()
        self.search_results_list.setMaximumHeight(120)
        self.search_results_list.setVisible(False)

        main_layout.addWidget(self.search_results_list)

        # =====================================================
        # Formatting Toolbar
        # =====================================================

        formatting_group = QGroupBox("Formatting")
        formatting_layout = QVBoxLayout()
        formatting_layout.setContentsMargins(10, 8, 10, 8)
        formatting_layout.setSpacing(8)

        # Row 1 - Font
        font_row = QHBoxLayout()
        font_row.setSpacing(6)

        self.bold_button = QPushButton("B")
        self.bold_button.setToolTip("Bold")
        self.bold_button.setCheckable(True)

        self.italic_button = QPushButton("I")
        self.italic_button.setToolTip("Italic")
        self.italic_button.setCheckable(True)

        self.underline_button = QPushButton("U")
        self.underline_button.setToolTip(
            "Underline. Click repeatedly for single/double/none."
        )
        self.underline_button.setCheckable(True)

        for button in (
            self.bold_button,
            self.italic_button,
            self.underline_button,
        ):
            button.setMinimumSize(38, 34)

        font_row.addWidget(self.bold_button)
        font_row.addWidget(self.italic_button)
        font_row.addWidget(self.underline_button)

        font_row.addWidget(QLabel("Font Size:"))

        self.font_size_spin = QSpinBox()
        self.font_size_spin.setRange(1, 72)
        self.font_size_spin.setValue(11)
        self.font_size_spin.setMinimumWidth(75)
        self.font_size_spin.setToolTip("Font size")

        font_row.addWidget(self.font_size_spin)

        font_row.addWidget(QLabel("Font Color:"))

        self.font_color_button = QPushButton("■")
        self.font_color_button.setMinimumSize(55, 34)
        self.font_color_button.setToolTip("Choose font color")
        self.set_color_button(self.font_color_button, "000000")

        font_row.addWidget(self.font_color_button)

        font_row.addWidget(QLabel("Fill Color:"))

        self.fill_color_button = QPushButton("■")
        self.fill_color_button.setMinimumSize(55, 34)
        self.fill_color_button.setToolTip("Choose fill color")
        self.set_color_button(self.fill_color_button, "FFFFFF")

        font_row.addWidget(self.fill_color_button)

        font_row.addStretch()

        formatting_layout.addLayout(font_row)

        # Row 2 - Alignment and number format
        format_row = QHBoxLayout()
        format_row.setSpacing(6)

        format_row.addWidget(QLabel("Alignment:"))

        self.left_align_button = QPushButton("Left")
        self.center_align_button = QPushButton("Center")
        self.right_align_button = QPushButton("Right")
        self.top_align_button = QPushButton("Top")
        self.middle_align_button = QPushButton("Middle")
        self.bottom_align_button = QPushButton("Bottom")

        for button in (
            self.left_align_button,
            self.center_align_button,
            self.right_align_button,
            self.top_align_button,
            self.middle_align_button,
            self.bottom_align_button,
        ):
            button.setMinimumHeight(34)

        format_row.addWidget(self.left_align_button)
        format_row.addWidget(self.center_align_button)
        format_row.addWidget(self.right_align_button)
        format_row.addSpacing(8)
        format_row.addWidget(self.top_align_button)
        format_row.addWidget(self.middle_align_button)
        format_row.addWidget(self.bottom_align_button)

        format_row.addSpacing(12)
        format_row.addWidget(QLabel("Number Format:"))

        self.number_format_combo = QComboBox()
        self.number_format_combo.addItems(
            [
                "General",
                "Number",
                "Currency",
                "Accounting",
                "Short Date",
                "Long Date",
                "Time",
                "Percentage",
                "Fraction",
                "Scientific",
                "Text",
            ]
        )
        self.number_format_combo.setMinimumWidth(150)
        self.number_format_combo.setMinimumHeight(34)

        format_row.addWidget(self.number_format_combo)
        format_row.addStretch()

        formatting_layout.addLayout(format_row)

        # Row 3 - Presets / advanced formatting
        action_row = QHBoxLayout()
        action_row.setSpacing(8)

        self.table_format_button = QPushButton("▦  Table Format")
        self.clear_format_button = QPushButton("🧹  Clear Format")
        self.format_cells_button = QPushButton("▦  Format Cells...")

        self.table_format_button.setMinimumHeight(36)
        self.clear_format_button.setMinimumHeight(36)
        self.format_cells_button.setMinimumHeight(36)

        action_row.addWidget(self.table_format_button)
        action_row.addWidget(self.clear_format_button)
        action_row.addStretch()
        action_row.addWidget(self.format_cells_button)

        formatting_layout.addLayout(action_row)

        formatting_group.setLayout(formatting_layout)
        main_layout.addWidget(formatting_group)

        # =====================================================
        # Table
        # =====================================================

        self.table = QTableWidget()

        self.table.setEditTriggers(
            QTableWidget.EditTrigger.DoubleClicked
            | QTableWidget.EditTrigger.EditKeyPressed
        )

        self.table.setAlternatingRowColors(True)

        self.table.setSelectionBehavior(
            QTableWidget.SelectionBehavior.SelectItems
        )

        self.table.cellChanged.connect(
            self.handle_cell_changed
        )
        self.table.itemSelectionChanged.connect(
            self.update_format_controls_from_selection
        )

        self.table.horizontalHeader().setStretchLastSection(
            True
        )

        self.search_button.clicked.connect(
            self.search_excel
        )
        self.clear_search_button.clicked.connect(
            self.clear_search
        )
        self.search_input.returnPressed.connect(
            self.search_excel
        )
        self.search_results_list.itemClicked.connect(
            self.go_to_search_result
        )

        main_layout.addWidget(self.table)

        # =====================================================
        # Bottom buttons
        # =====================================================

        bottom_layout = QHBoxLayout()

        self.add_row_button = QPushButton("Add Row")
        self.delete_row_button = QPushButton("Delete Row")
        self.add_column_button = QPushButton("Add Column")
        self.rename_column_button = QPushButton("Rename Column")
        self.delete_column_button = QPushButton("Delete Column")
        self.save_button = QPushButton("Save")

        for button in (
            self.add_row_button,
            self.delete_row_button,
            self.add_column_button,
            self.rename_column_button,
            self.delete_column_button,
            self.save_button,
        ):
            button.setMinimumHeight(40)

        bottom_layout.addWidget(self.add_row_button)
        bottom_layout.addWidget(self.delete_row_button)
        bottom_layout.addWidget(self.add_column_button)
        bottom_layout.addWidget(self.rename_column_button)
        bottom_layout.addWidget(self.delete_column_button)
        bottom_layout.addStretch()
        bottom_layout.addWidget(self.save_button)

        main_layout.addLayout(bottom_layout)

        self.setLayout(main_layout)

    # =========================================================
    # Helpers
    # =========================================================

    @staticmethod
    def column_letter(column_number: int) -> str:
        result = ""

        while column_number > 0:
            column_number, remainder = divmod(
                column_number - 1,
                26,
            )
            result = (
                chr(65 + remainder)
                + result
            )

        return result

    def selected_sheet(self) -> str:
        return self.sheet_combo.currentText().strip()

    def set_busy(self, button, text: str, busy: bool):
        button.setEnabled(not busy)
        button.setText(text if busy else button.property("normal_text"))

    def clear_pending_for_sheet(self, sheet_name: str):
        self.pending_changes = {
            key: value
            for key, value in self.pending_changes.items()
            if key[0] != sheet_name
        }

    def rebuild_sheet_combo(
        self,
        selected_sheet: str | None = None,
    ):
        if not self.workbook:
            return

        names = self.workbook.sheetnames

        self.sheet_combo.blockSignals(True)
        try:
            self.sheet_combo.clear()
            self.sheet_combo.addItems(names)

            if selected_sheet in names:
                self.sheet_combo.setCurrentText(selected_sheet)
            elif names:
                self.sheet_combo.setCurrentIndex(0)
        finally:
            self.sheet_combo.blockSignals(False)

    # =========================================================
    # Download + Load Workbook
    # =========================================================

    def load_excel_file(self):
        try:
            self.refresh_button.setEnabled(False)
            self.refresh_button.setText("Loading...")

            self.pending_changes.clear()

            if self.temp_file_path:
                try:
                    os.remove(self.temp_file_path)
                except OSError:
                    pass

            fd, self.temp_file_path = tempfile.mkstemp(
                suffix=".xlsx"
            )
            os.close(fd)

            self.api_client.download_file(
                file_id=self.file_id,
                save_path=self.temp_file_path,
            )

            self.workbook = load_workbook(
                self.temp_file_path
            )

            current_sheet = self.selected_sheet()

            self.rebuild_sheet_combo(
                current_sheet
                if current_sheet in self.workbook.sheetnames
                else None
            )

            if self.workbook.sheetnames:
                self.load_sheet(
                    self.sheet_combo.currentText()
                )
            else:
                self.table.clear()
                self.table.setRowCount(0)
                self.table.setColumnCount(0)

        except Exception as exc:
            QMessageBox.critical(
                self,
                "Failed to Load Excel File",
                str(exc),
            )

        finally:
            self.refresh_button.setEnabled(True)
            self.refresh_button.setText("Refresh")

    # =========================================================
    # Load Sheet
    # =========================================================

    def load_sheet(self, sheet_name: str):
        if not self.workbook or not sheet_name:
            return

        if sheet_name not in self.workbook.sheetnames:
            return

        self._loading_table = True
        self.table.blockSignals(True)
        self.clear_search()

        try:
            worksheet = self.workbook[sheet_name]

            max_row = worksheet.max_row
            max_column = worksheet.max_column

            # A newly created completely empty sheet has
            # max_row/max_column equal to 1 in some openpyxl
            # situations. We still show one editable cell.
            if max_row < 1:
                max_row = 1

            if max_column < 1:
                max_column = 1

            self.table.clear()
            self.table.setRowCount(max_row)
            self.table.setColumnCount(max_column)

            headers = [
                self.column_letter(index + 1)
                for index in range(max_column)
            ]

            self.table.setHorizontalHeaderLabels(headers)

            for row_index in range(max_row):
                excel_row = row_index + 1

                for column_index in range(max_column):
                    excel_column = column_index + 1

                    value = worksheet.cell(
                        row=excel_row,
                        column=excel_column,
                    ).value

                    text = "" if value is None else str(value)

                    item = QTableWidgetItem(text)

                    # Restore workbook formatting in the desktop preview.
                    self.apply_openpyxl_style_to_item(
                        item,
                        worksheet.cell(
                            row=excel_row,
                            column=excel_column,
                        ),
                    )

                    self.table.setItem(
                        row_index,
                        column_index,
                        item,
                    )

            self.table.resizeColumnsToContents()

        except Exception as exc:
            QMessageBox.critical(
                self,
                "Failed to Load Sheet",
                str(exc),
            )

        finally:
            self.table.blockSignals(False)
            self._loading_table = False

    # =========================================================
    # Excel Search
    # =========================================================

    def search_excel(self):
        sheet_name = self.selected_sheet()
        search_term = self.search_input.text().strip()

        if not sheet_name:
            QMessageBox.warning(
                self,
                "Search Excel",
                "Please select a sheet first.",
            )
            return

        if not search_term:
            QMessageBox.warning(
                self,
                "Search Excel",
                "Please enter a search term.",
            )
            return

        self.search_button.setEnabled(False)
        self.search_button.setText("Searching...")

        try:
            result = self.api_client.search_excel(
                file_id=self.file_id,
                sheet_name=sheet_name,
                search_term=search_term,
            )

            results = result.get("results", [])
            self.search_results = results

            self.search_results_list.clear()

            if not results:
                self.search_results_list.setVisible(False)
                self.search_result_label.setText(
                    "0 results"
                )

                QMessageBox.information(
                    self,
                    "Search Excel",
                    (
                        f"No matches found for "
                        f"'{search_term}'."
                    ),
                )
                return

            for index, match in enumerate(results):
                cell = match.get("cell", "")
                value = match.get("value", "")
                text = (
                    f"{cell}  |  "
                    f"{value}"
                )

                item = QListWidgetItem(text)
                item.setData(
                    0x0100,
                    index,
                )

                self.search_results_list.addItem(item)

            self.search_results_list.setVisible(True)
            self.search_result_label.setText(
                f"{len(results)} result(s)"
            )

            # Automatically select the first result.
            self.go_to_search_result(
                self.search_results_list.item(0)
            )

        except Exception as exc:
            self.search_results = []
            self.search_results_list.clear()
            self.search_results_list.setVisible(False)
            self.search_result_label.setText(
                "Search failed."
            )

            QMessageBox.critical(
                self,
                "Search Failed",
                str(exc),
            )

        finally:
            self.search_button.setEnabled(True)
            self.search_button.setText("Search")

    def go_to_search_result(
        self,
        item: QListWidgetItem | None,
    ):
        if item is None:
            return

        index = item.data(0x0100)

        if index is None:
            return

        try:
            index = int(index)
            result = self.search_results[index]
        except (IndexError, TypeError, ValueError):
            return

        cell = result.get("cell")

        if not cell:
            return

        row_number = result.get("row_number")
        column_number = result.get("column_number")

        # Prefer the explicit row/column returned by the backend.
        try:
            if row_number is not None and column_number is not None:
                row = int(row_number) - 1
                column = int(column_number) - 1
            else:
                row, column = self.cell_to_indexes(cell)
        except (TypeError, ValueError):
            return

        if (
            row < 0
            or column < 0
            or row >= self.table.rowCount()
            or column >= self.table.columnCount()
        ):
            return

        self.table.setCurrentCell(row, column)
        self.table.scrollToItem(
            self.table.item(row, column)
        )
        self.table.setFocus()

    @staticmethod
    def cell_to_indexes(cell: str) -> tuple[int, int]:
        letters = ""
        digits = ""

        for char in cell.upper():
            if char.isalpha():
                letters += char
            elif char.isdigit():
                digits += char

        if not letters or not digits:
            raise ValueError(
                f"Invalid cell reference: {cell}"
            )

        column_number = 0

        for char in letters:
            column_number = (
                column_number * 26
                + ord(char)
                - ord("A")
                + 1
            )

        row_number = int(digits)

        return row_number - 1, column_number - 1

    def clear_search(self):
        self.search_input.clear()
        self.search_results.clear()
        self.search_results_list.clear()
        self.search_results_list.setVisible(False)
        self.search_result_label.setText(
            "No search performed."
        )

    # =========================================================
    # Formatting Helpers
    # =========================================================

    @staticmethod
    def normalize_color(color: str | None) -> str | None:
        if not color:
            return None

        value = str(color).strip().replace("#", "").upper()

        # openpyxl may return ARGB (8 characters).
        if len(value) == 8:
            value = value[-6:]

        if len(value) != 6:
            return None

        try:
            int(value, 16)
        except ValueError:
            return None

        return value

    @staticmethod
    def set_color_button(button: QPushButton, color: str):
        color = color.replace("#", "").upper()
        button.setStyleSheet(
            f"""
            QPushButton {{
                background-color: #{color};
                border: 1px solid #999999;
                border-radius: 4px;
                font-size: 18px;
            }}
            QPushButton:hover {{
                border: 2px solid #333333;
            }}
            """
        )
        button.setProperty("color_hex", color)

    def selected_range(self) -> str | None:
        ranges = self.table.selectedRanges()

        if not ranges:
            return None

        selection = ranges[0]

        start = (
            self.column_letter(selection.leftColumn() + 1)
            + str(selection.topRow() + 1)
        )

        end = (
            self.column_letter(selection.rightColumn() + 1)
            + str(selection.bottomRow() + 1)
        )

        return start if start == end else f"{start}:{end}"

    def selected_cells(self):
        ranges = self.table.selectedRanges()

        if not ranges:
            return []

        selection = ranges[0]

        cells = []

        for row in range(
            selection.topRow(),
            selection.bottomRow() + 1,
        ):
            for column in range(
                selection.leftColumn(),
                selection.rightColumn() + 1,
            ):
                cell = (
                    self.column_letter(column + 1)
                    + str(row + 1)
                )
                cells.append(cell)

        return cells

    def format_request(
        self,
        *,
        bold=None,
        italic=None,
        underline=None,
        font_size=None,
        font_color=None,
        fill_color=None,
        horizontal_alignment=None,
        vertical_alignment=None,
        number_format=None,
    ):
        sheet_name = self.selected_sheet()
        cell_range = self.selected_range()

        if not sheet_name:
            QMessageBox.warning(
                self,
                "Formatting",
                "Please select a sheet first.",
            )
            return False

        if not cell_range:
            QMessageBox.warning(
                self,
                "Formatting",
                "Please select one or more cells first.",
            )
            return False

        if self.pending_changes:
            QMessageBox.warning(
                self,
                "Formatting",
                (
                    "Please save your cell changes before "
                    "applying formatting."
                ),
            )
            return False

        selected_cells = self.selected_cells()

        if not selected_cells:
            QMessageBox.warning(
                self,
                "Formatting",
                "Please select one or more cells first.",
            )
            return False

        try:
            # -----------------------------------------------------
            # IMPORTANT: update the local preview FIRST.
            # This makes the formatting visible immediately in the
            # QTableWidget instead of waiting for the API response.
            # -----------------------------------------------------
            self.apply_local_format(
                sheet_name=sheet_name,
                cells=selected_cells,
                bold=bold,
                italic=italic,
                underline=underline,
                font_size=font_size,
                font_color=font_color,
                fill_color=fill_color,
                horizontal_alignment=horizontal_alignment,
                vertical_alignment=vertical_alignment,
                number_format=number_format,
            )

            # Force Qt to paint the changed cells before the synchronous
            # HTTP request starts. Without this, the UI may appear to
            # change only after the API request finishes.
            QApplication.processEvents()

            # -----------------------------------------------------
            # Persist the same formatting in the backend.
            # -----------------------------------------------------
            self.api_client.format_excel_range(
                file_id=self.file_id,
                sheet_name=sheet_name,
                cell_range=cell_range,
                bold=bold,
                italic=italic,
                underline=underline,
                font_size=font_size,
                font_color=font_color,
                fill_color=fill_color,
                horizontal_alignment=horizontal_alignment,
                vertical_alignment=vertical_alignment,
                number_format=number_format,
            )

            self.refresh_format_controls()
            return True

        except Exception as exc:
            # The local preview was already changed. If the backend
            # failed, reload the workbook so the UI returns to the
            # server's actual state. Formatting is blocked while there
            # are pending cell edits, so this is safe here.
            try:
                self.load_excel_file()
            except Exception:
                pass

            QMessageBox.critical(
                self,
                "Formatting Failed",
                str(exc),
            )
            return False

    def apply_local_format(
        self,
        *,
        sheet_name,
        cells,
        bold=None,
        italic=None,
        underline=None,
        font_size=None,
        font_color=None,
        fill_color=None,
        horizontal_alignment=None,
        vertical_alignment=None,
        number_format=None,
    ):
        worksheet = self.workbook[sheet_name]

        for cell in cells:
            ws_cell = worksheet[cell]

            if bold is not None:
                font = copy(ws_cell.font)
                font.bold = bold
                ws_cell.font = font

            if italic is not None:
                font = copy(ws_cell.font)
                font.italic = italic
                ws_cell.font = font

            if underline is not None:
                font = copy(ws_cell.font)
                font.underline = underline
                ws_cell.font = font

            if font_size is not None:
                font = copy(ws_cell.font)
                font.sz = font_size
                ws_cell.font = font

            normalized_font = self.normalize_color(font_color)
            if normalized_font:
                font = copy(ws_cell.font)
                font.color = normalized_font
                ws_cell.font = font

            normalized_fill = self.normalize_color(fill_color)
            if normalized_fill:
                from openpyxl.styles import PatternFill

                ws_cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=normalized_fill,
                )

            if horizontal_alignment is not None:
                alignment = copy(ws_cell.alignment)
                alignment.horizontal = horizontal_alignment
                ws_cell.alignment = alignment

            if vertical_alignment is not None:
                alignment = copy(ws_cell.alignment)
                alignment.vertical = vertical_alignment
                ws_cell.alignment = alignment

            if number_format is not None:
                ws_cell.number_format = number_format

            row = ws_cell.row - 1
            column = ws_cell.column - 1

            item = self.table.item(row, column)
            if item is not None:
                self.apply_openpyxl_style_to_item(
                    item,
                    ws_cell,
                )

    @staticmethod
    def apply_openpyxl_style_to_item(item, ws_cell):
        font = ws_cell.font
        alignment = ws_cell.alignment
        fill = ws_cell.fill

        qfont = QFont()
        qfont.setBold(bool(font.bold))
        qfont.setItalic(bool(font.italic))
        qfont.setPointSize(
            max(1, int(font.sz or 11))
        )

        if font.underline in {
            "single",
            "singleAccounting",
        }:
            qfont.setUnderline(True)
            qfont.setStrikeOut(False)
        elif font.underline in {
            "double",
            "doubleAccounting",
        }:
            qfont.setUnderline(True)
            qfont.setStrikeOut(False)

        item.setFont(qfont)

        if font.color and font.color.type == "rgb":
            rgb = font.color.rgb
            if rgb:
                normalized = str(rgb)[-6:]
                item.setForeground(
                    QColor(f"#{normalized}")
                )

        if fill.fill_type == "solid":
            fg = fill.fgColor
            if fg.type == "rgb" and fg.rgb:
                normalized = str(fg.rgb)[-6:]
                item.setBackground(
                    QColor(f"#{normalized}")
                )

        alignment_map = {
            "left": Qt.AlignmentFlag.AlignLeft,
            "center": Qt.AlignmentFlag.AlignHCenter,
            "right": Qt.AlignmentFlag.AlignRight,
        }

        vertical_map = {
            "top": Qt.AlignmentFlag.AlignTop,
            "center": Qt.AlignmentFlag.AlignVCenter,
            "bottom": Qt.AlignmentFlag.AlignBottom,
        }

        horizontal_flag = alignment_map.get(
            alignment.horizontal,
            Qt.AlignmentFlag.AlignLeft,
        )
        vertical_flag = vertical_map.get(
            alignment.vertical,
            Qt.AlignmentFlag.AlignVCenter,
        )

        qt_alignment = horizontal_flag | vertical_flag

        item.setTextAlignment(qt_alignment)

    def refresh_format_controls(self):
        self.table.blockSignals(True)
        try:
            self.update_format_controls_from_selection()
        finally:
            self.table.blockSignals(False)

    def update_format_controls_from_selection(self):
        ranges = self.table.selectedRanges()

        if not ranges or not self.workbook:
            return

        selection = ranges[0]
        sheet_name = self.selected_sheet()

        if not sheet_name:
            return

        worksheet = self.workbook[sheet_name]
        cell = worksheet.cell(
            row=selection.topRow() + 1,
            column=selection.leftColumn() + 1,
        )

        self.bold_button.blockSignals(True)
        self.italic_button.blockSignals(True)
        self.underline_button.blockSignals(True)
        self.font_size_spin.blockSignals(True)
        self.number_format_combo.blockSignals(True)

        try:
            self.bold_button.setChecked(
                bool(cell.font.bold)
            )
            self.italic_button.setChecked(
                bool(cell.font.italic)
            )

            underline = cell.font.underline
            self.underline_style = underline

            self.underline_button.setChecked(
                underline is not None
            )

            self.font_size_spin.setValue(
                max(1, int(cell.font.sz or 11))
            )

            format_name = self.number_format_to_name(
                cell.number_format
            )

            index = self.number_format_combo.findText(
                format_name
            )

            if index >= 0:
                self.number_format_combo.setCurrentIndex(
                    index
                )

            if (
                cell.font.color
                and cell.font.color.type == "rgb"
                and cell.font.color.rgb
            ):
                self.set_color_button(
                    self.font_color_button,
                    str(cell.font.color.rgb)[-6:],
                )

            if (
                cell.fill.fill_type == "solid"
                and cell.fill.fgColor.type == "rgb"
                and cell.fill.fgColor.rgb
            ):
                self.set_color_button(
                    self.fill_color_button,
                    str(cell.fill.fgColor.rgb)[-6:],
                )

        finally:
            self.bold_button.blockSignals(False)
            self.italic_button.blockSignals(False)
            self.underline_button.blockSignals(False)
            self.font_size_spin.blockSignals(False)
            self.number_format_combo.blockSignals(False)

    @staticmethod
    def number_format_to_name(number_format: str) -> str:
        if number_format in {
            "General",
            "@",
        }:
            return (
                "General"
                if number_format == "General"
                else "Text"
            )

        mapping = {
            "0": "Number",
            "0.00": "Number",
            "#,##0": "Number",
            "#,##0.00": "Number",
            '$#,##0.00': "Currency",
            '$#,##0.00_);($#,##0.00)': "Currency",
            '_-* #,##0.00\\ [$£-809]_-;\\-* #,##0.00\\ [$£-809]_-;_-* "-"??\\ [$£-809]_-;_-@_-': "Accounting",
            "m/d/yy": "Short Date",
            "mm/dd/yy": "Short Date",
            "dd/mm/yy": "Short Date",
            "dddd, mmmm d, yyyy": "Long Date",
            "h:mm AM/PM": "Time",
            "0%": "Percentage",
            "0.00%": "Percentage",
            "# ?/?": "Fraction",
            "0.00E+00": "Scientific",
        }

        return mapping.get(number_format, "General")

    def toggle_bold(self):
        if self._loading_table:
            return
        self.format_request(
            bold=self.bold_button.isChecked()
        )

    def toggle_italic(self):
        if self._loading_table:
            return
        self.format_request(
            italic=self.italic_button.isChecked()
        )

    def toggle_underline(self):
        current = self.underline_style

        if current is None:
            new_style = "single"
        elif current == "single":
            new_style = "double"
        else:
            new_style = None

        self.underline_style = new_style

        success = self.format_request(
            underline=new_style
        )

        if not success:
            self.underline_style = current

    def change_font_size(self, value: int):
        if self._loading_table:
            return
        self.format_request(
            font_size=float(value)
        )

    def choose_font_color(self):
        current = self.font_color_button.property(
            "color_hex"
        ) or "000000"

        color = QColorDialog.getColor(
            QColor(f"#{current}"),
            self,
            "Choose Font Color",
        )

        if not color.isValid():
            return

        value = color.name().replace("#", "").upper()

        if self.format_request(font_color=value):
            self.set_color_button(
                self.font_color_button,
                value,
            )

    def choose_fill_color(self):
        current = self.fill_color_button.property(
            "color_hex"
        ) or "FFFFFF"

        color = QColorDialog.getColor(
            QColor(f"#{current}"),
            self,
            "Choose Fill Color",
        )

        if not color.isValid():
            return

        value = color.name().replace("#", "").upper()

        if self.format_request(fill_color=value):
            self.set_color_button(
                self.fill_color_button,
                value,
            )

    def apply_alignment(self, alignment: str):
        self.format_request(
            horizontal_alignment=alignment
        )

    def apply_vertical_alignment(self, alignment: str):
        self.format_request(
            vertical_alignment=alignment
        )

    def change_number_format(self, format_name: str):
        if self._loading_table:
            return

        formats = {
            "General": "General",
            "Number": "0.00",
            "Currency": '$#,##0.00',
            "Accounting": '_-* #,##0.00\\ [$£-809]_-;\\-* #,##0.00\\ [$£-809]_-;_-* "-"??\\ [$£-809]_-;_-@_-',
            "Short Date": "mm/dd/yy",
            "Long Date": "dddd, mmmm d, yyyy",
            "Time": "h:mm AM/PM",
            "Percentage": "0.00%",
            "Fraction": "# ?/?",
            "Scientific": "0.00E+00",
            "Text": "@",
        }

        number_format = formats.get(format_name)

        if number_format is None:
            return

        self.format_request(
            number_format=number_format
        )

    def apply_table_format(self):
        sheet_name = self.selected_sheet()
        cell_range = self.selected_range()

        if not sheet_name or not cell_range:
            QMessageBox.warning(
                self,
                "Table Format",
                "Please select a range first.",
            )
            return

        if self.pending_changes:
            QMessageBox.warning(
                self,
                "Table Format",
                (
                    "Please save your cell changes before "
                    "applying formatting."
                ),
            )
            return

        ranges = self.table.selectedRanges()
        selection = ranges[0]

        start_row = selection.topRow() + 1
        end_row = selection.bottomRow() + 1
        start_column = selection.leftColumn() + 1
        end_column = selection.rightColumn() + 1

        header_start = (
            self.column_letter(start_column)
            + str(start_row)
        )
        header_end = (
            self.column_letter(end_column)
            + str(start_row)
        )
        header_range = (
            header_start
            if header_start == header_end
            else f"{header_start}:{header_end}"
        )

        try:
            # The current backend has a generic format endpoint,
            # not a dedicated Excel Table endpoint. Therefore this
            # preset creates a practical table-like appearance by
            # formatting the header and data area.
            self.api_client.format_excel_range(
                file_id=self.file_id,
                sheet_name=sheet_name,
                cell_range=header_range,
                bold=True,
                italic=False,
                underline=None,
                font_size=11,
                font_color="FFFFFF",
                fill_color="4472C4",
                horizontal_alignment="center",
                vertical_alignment="center",
                number_format=None,
            )

            body_start = (
                self.column_letter(start_column)
                + str(start_row + 1)
            )
            body_end = (
                self.column_letter(end_column)
                + str(end_row)
            )

            if start_row < end_row:
                body_range = f"{body_start}:{body_end}"

                self.api_client.format_excel_range(
                    file_id=self.file_id,
                    sheet_name=sheet_name,
                    cell_range=body_range,
                    bold=False,
                    italic=False,
                    underline=None,
                    font_size=11,
                    font_color="000000",
                    fill_color="FFFFFF",
                    horizontal_alignment="left",
                    vertical_alignment="center",
                    number_format=None,
                )

            # Update the local preview from the same style.
            worksheet = self.workbook[sheet_name]

            for column in range(
                start_column,
                end_column + 1,
            ):
                cell = worksheet.cell(
                    row=start_row,
                    column=column,
                )
                font = copy(cell.font)
                font.bold = True
                font.color = "FFFFFF"
                font.sz = 11
                cell.font = font
                from openpyxl.styles import PatternFill
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="4472C4",
                )
                alignment = copy(cell.alignment)
                alignment.horizontal = "center"
                alignment.vertical = "center"
                cell.alignment = alignment
                self.apply_openpyxl_style_to_item(
                    self.table.item(
                        start_row - 1,
                        column - 1,
                    ),
                    cell,
                )

            for row in range(
                start_row + 1,
                end_row + 1,
            ):
                for column in range(
                    start_column,
                    end_column + 1,
                ):
                    cell = worksheet.cell(
                        row=row,
                        column=column,
                    )
                    font = copy(cell.font)
                    font.bold = False
                    font.italic = False
                    font.color = "000000"
                    font.sz = 11
                    cell.font = font
                    cell.fill = PatternFill(
                        fill_type="solid",
                        fgColor="FFFFFF",
                    )
                    alignment = copy(cell.alignment)
                    alignment.horizontal = "left"
                    alignment.vertical = "center"
                    cell.alignment = alignment
                    self.apply_openpyxl_style_to_item(
                        self.table.item(
                            row - 1,
                            column - 1,
                        ),
                        cell,
                    )

            QMessageBox.information(
                self,
                "Table Format",
                "Table formatting applied successfully.",
            )

        except Exception as exc:
            QMessageBox.critical(
                self,
                "Table Format Failed",
                str(exc),
            )

    def clear_format(self):
        if not self.selected_range():
            QMessageBox.warning(
                self,
                "Clear Format",
                "Please select one or more cells first.",
            )
            return

        if self.pending_changes:
            QMessageBox.warning(
                self,
                "Clear Format",
                (
                    "Please save your cell changes before "
                    "clearing formatting."
                ),
            )
            return

        # Reset using the generic backend formatting endpoint.
        # The backend does not expose a separate clear-format API,
        # so we explicitly reset the supported properties.
        success = self.format_request(
            bold=False,
            italic=False,
            underline=None,
            font_size=11,
            font_color="000000",
            fill_color="FFFFFF",
            horizontal_alignment="left",
            vertical_alignment="center",
            number_format="General",
        )

        if success:
            self.underline_style = None

    def open_format_cells_dialog(self):
        if not self.selected_range():
            QMessageBox.warning(
                self,
                "Format Cells",
                "Please select one or more cells first.",
            )
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Format Cells")
        dialog.setMinimumSize(560, 420)

        tabs = QTabWidget()

        # Font tab
        font_tab = QWidget()
        font_form = QFormLayout()

        font_size = QSpinBox()
        font_size.setRange(1, 72)
        font_size.setValue(self.font_size_spin.value())

        font_bold = QPushButton("Bold")
        font_bold.setCheckable(True)
        font_bold.setChecked(
            self.bold_button.isChecked()
        )

        font_italic = QPushButton("Italic")
        font_italic.setCheckable(True)
        font_italic.setChecked(
            self.italic_button.isChecked()
        )

        font_underline = QComboBox()
        font_underline.addItems(
            [
                "None",
                "single",
                "double",
                "singleAccounting",
                "doubleAccounting",
            ]
        )

        current_underline = self.underline_style
        if current_underline:
            index = font_underline.findText(
                current_underline
            )
            if index >= 0:
                font_underline.setCurrentIndex(index)

        font_form.addRow("Font Size:", font_size)
        font_form.addRow("Bold:", font_bold)
        font_form.addRow("Italic:", font_italic)
        font_form.addRow(
            "Underline:",
            font_underline,
        )

        font_tab.setLayout(font_form)
        tabs.addTab(font_tab, "Font")

        # Alignment tab
        alignment_tab = QWidget()
        alignment_form = QFormLayout()

        horizontal = QComboBox()
        horizontal.addItems(
            ["left", "center", "right"]
        )

        vertical = QComboBox()
        vertical.addItems(
            ["top", "center", "bottom"]
        )

        alignment_form.addRow(
            "Horizontal:",
            horizontal,
        )
        alignment_form.addRow(
            "Vertical:",
            vertical,
        )

        alignment_tab.setLayout(alignment_form)
        tabs.addTab(alignment_tab, "Alignment")

        # Number tab
        number_tab = QWidget()
        number_form = QFormLayout()

        number_combo = QComboBox()
        number_combo.addItems(
            [
                "General",
                "Number",
                "Currency",
                "Accounting",
                "Short Date",
                "Long Date",
                "Time",
                "Percentage",
                "Fraction",
                "Scientific",
                "Text",
            ]
        )
        number_combo.setCurrentText(
            self.number_format_combo.currentText()
        )

        number_form.addRow(
            "Category:",
            number_combo,
        )

        number_tab.setLayout(number_form)
        tabs.addTab(number_tab, "Number")

        # Fill tab
        fill_tab = QWidget()
        fill_form = QFormLayout()

        fill_button = QPushButton("Choose Fill Color")
        fill_button.clicked.connect(
            lambda: self.choose_dialog_color(
                fill_button
            )
        )

        fill_form.addRow(
            "Fill:",
            fill_button,
        )

        fill_tab.setLayout(fill_form)
        tabs.addTab(fill_tab, "Fill")

        # Buttons
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok
            | QDialogButtonBox.StandardButton.Cancel
        )

        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)

        layout = QVBoxLayout()
        layout.addWidget(tabs)
        layout.addWidget(buttons)
        dialog.setLayout(layout)

        if dialog.exec() != QDialog.DialogCode.Accepted:
            return

        underline_value = font_underline.currentText()
        if underline_value == "None":
            underline_value = None

        formats = {
            "General": "General",
            "Number": "0.00",
            "Currency": '$#,##0.00',
            "Accounting": '_-* #,##0.00\\ [$£-809]_-;\\-* #,##0.00\\ [$£-809]_-;_-* "-"??\\ [$£-809]_-;_-@_-',
            "Short Date": "mm/dd/yy",
            "Long Date": "dddd, mmmm d, yyyy",
            "Time": "h:mm AM/PM",
            "Percentage": "0.00%",
            "Fraction": "# ?/?",
            "Scientific": "0.00E+00",
            "Text": "@",
        }

        success = self.format_request(
            bold=font_bold.isChecked(),
            italic=font_italic.isChecked(),
            underline=underline_value,
            font_size=float(font_size.value()),
            horizontal_alignment=horizontal.currentText(),
            vertical_alignment=vertical.currentText(),
            number_format=formats[
                number_combo.currentText()
            ],
        )

        if success:
            self.underline_style = underline_value

    def choose_dialog_color(self, button: QPushButton):
        color = QColorDialog.getColor(
            QColor("#FFFFFF"),
            self,
            "Choose Fill Color",
        )

        if color.isValid():
            value = color.name().replace(
                "#", ""
            ).upper()

            button.setText(
                f"#{value}"
            )
            button.setProperty(
                "color_hex",
                value,
            )

            # Apply immediately.
            self.format_request(
                fill_color=value
            )

    # =========================================================
    # Cell Editing
    # =========================================================

    def handle_cell_changed(
        self,
        row: int,
        column: int,
    ):
        if self._loading_table:
            return

        if not self.workbook:
            return

        sheet_name = self.selected_sheet()

        if not sheet_name:
            return

        item = self.table.item(row, column)

        value = None if item is None else item.text()

        cell = (
            self.column_letter(column + 1)
            + str(row + 1)
        )

        self.pending_changes[
            (sheet_name, cell)
        ] = value

        try:
            worksheet = self.workbook[sheet_name]
            worksheet[cell] = value
        except Exception as exc:
            QMessageBox.critical(
                self,
                "Failed to Update Cell",
                str(exc),
            )

    # =========================================================
    # Save Cell Changes
    # =========================================================

    def save_changes(self):
        if not self.pending_changes:
            QMessageBox.information(
                self,
                "Save",
                "There are no changes to save.",
            )
            return

        self.save_button.setEnabled(False)
        self.save_button.setText("Saving...")

        try:
            successful_changes = []

            for (
                sheet_name,
                cell,
            ), value in list(
                self.pending_changes.items()
            ):
                self.api_client.update_cell(
                    file_id=self.file_id,
                    sheet_name=sheet_name,
                    cell=cell,
                    value=value,
                )

                successful_changes.append(
                    (sheet_name, cell)
                )

            for change in successful_changes:
                self.pending_changes.pop(
                    change,
                    None,
                )

            QMessageBox.information(
                self,
                "Save Successful",
                "Changes saved successfully.",
            )

        except Exception as exc:
            QMessageBox.critical(
                self,
                "Save Failed",
                (
                    "Some changes could not be saved.\n\n"
                    f"{exc}"
                ),
            )

        finally:
            self.save_button.setEnabled(True)
            self.save_button.setText("Save")

    # =========================================================
    # Add Row
    # =========================================================

    def add_row(self):
        sheet_name = self.selected_sheet()

        if not sheet_name:
            QMessageBox.warning(
                self,
                "Add Row",
                "Please select a sheet first.",
            )
            return

        column_count = max(
            self.table.columnCount(),
            1,
        )

        # IMPORTANT:
        # Use empty strings instead of a list containing only None.
        # This makes the blank row a real row in openpyxl and prevents
        # max_row from collapsing back to the previous row.
        row_data = [""] * column_count

        self.add_row_button.setEnabled(False)
        self.add_row_button.setText("Adding...")

        try:
            result = self.api_client.add_row(
                file_id=self.file_id,
                sheet_name=sheet_name,
                row_data=row_data,
            )

            # The backend returns the actual Excel row number.
            # Use it instead of assuming table.rowCount() is the
            # correct Excel row.
            backend_row_number = result.get("row_number")

            if backend_row_number is None:
                backend_row_number = (
                    self.table.rowCount() + 1
                )

            # Update local workbook first.
            worksheet = self.workbook[sheet_name]

            # The backend has already added the row. The local
            # workbook is only a UI-side copy, so append one row.
            worksheet.append(row_data)

            # If openpyxl's calculated row differs from the backend
            # row, explicitly write the cells at the backend row.
            # This keeps the local workbook aligned with the server.
            local_row_number = worksheet.max_row

            if local_row_number != backend_row_number:
                for column_number, value in enumerate(
                    row_data,
                    start=1,
                ):
                    worksheet.cell(
                        row=backend_row_number,
                        column=column_number,
                        value=value,
                    )

            # Update the table without triggering cellChanged.
            self._loading_table = True
            self.table.blockSignals(True)

            try:
                target_row = backend_row_number - 1

                while self.table.rowCount() <= target_row:
                    self.table.insertRow(
                        self.table.rowCount()
                    )

                for column in range(column_count):
                    self.table.setItem(
                        target_row,
                        column,
                        QTableWidgetItem(""),
                    )

            finally:
                self.table.blockSignals(False)
                self._loading_table = False

            self.table.resizeColumnsToContents()
            self.table.scrollToBottom()

            QMessageBox.information(
                self,
                "Row Added",
                result.get(
                    "message",
                    "Row added successfully.",
                ),
            )

        except Exception as exc:
            QMessageBox.critical(
                self,
                "Add Row Failed",
                str(exc),
            )

        finally:
            self.add_row_button.setEnabled(True)
            self.add_row_button.setText("Add Row")

    # =========================================================
    # Delete Row
    # =========================================================

    def delete_row(self):
        current_row = self.table.currentRow()

        if current_row < 0:
            QMessageBox.warning(
                self,
                "Delete Row",
                "Please select a row to delete.",
            )
            return

        sheet_name = self.selected_sheet()

        if not sheet_name:
            QMessageBox.warning(
                self,
                "Delete Row",
                "Please select a sheet first.",
            )
            return

        # QTableWidget is 0-based.
        # Excel/openpyxl is 1-based.
        row_number = current_row + 1

        answer = QMessageBox.question(
            self,
            "Delete Row",
            (
                f"Are you sure you want to delete "
                f"row {row_number}?"
            ),
            QMessageBox.StandardButton.Yes
            | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )

        if answer != QMessageBox.StandardButton.Yes:
            return

        self.delete_row_button.setEnabled(False)
        self.delete_row_button.setText("Deleting...")

        try:
            result = self.api_client.delete_row(
                file_id=self.file_id,
                sheet_name=sheet_name,
                row_number=row_number,
            )

            # Backend succeeded. Now update the local workbook/table.
            worksheet = self.workbook[sheet_name]

            worksheet.delete_rows(
                row_number,
                1,
            )

            self._loading_table = True
            self.table.blockSignals(True)

            try:
                self.table.removeRow(current_row)
            finally:
                self.table.blockSignals(False)
                self._loading_table = False

            # Row deletion changes Excel coordinates for all rows below
            # the deleted row. Pending cell changes must therefore be
            # rebuilt with shifted row numbers.
            self._shift_pending_changes_after_row_delete(
                sheet_name,
                row_number,
            )

            QMessageBox.information(
                self,
                "Row Deleted",
                result.get(
                    "message",
                    "Row deleted successfully.",
                ),
            )

        except Exception as exc:
            QMessageBox.critical(
                self,
                "Delete Row Failed",
                str(exc),
            )

        finally:
            self.delete_row_button.setEnabled(True)
            self.delete_row_button.setText("Delete Row")

    def _shift_pending_changes_after_row_delete(
        self,
        sheet_name: str,
        deleted_row: int,
    ):
        updated = {}

        for (
            change_sheet,
            cell,
        ), value in self.pending_changes.items():

            if change_sheet != sheet_name:
                updated[(change_sheet, cell)] = value
                continue

            letters = ""
            digits = ""

            for char in cell:
                if char.isalpha():
                    letters += char
                elif char.isdigit():
                    digits += char

            if not digits:
                updated[(change_sheet, cell)] = value
                continue

            row_number = int(digits)

            if row_number == deleted_row:
                # This pending change belonged to the deleted row.
                continue

            if row_number > deleted_row:
                row_number -= 1

            updated[
                (
                    change_sheet,
                    f"{letters}{row_number}",
                )
            ] = value

        self.pending_changes = updated

    # =========================================================
    # Add Column
    # =========================================================

    def add_column(self):
        sheet_name = self.selected_sheet()

        if not sheet_name:
            QMessageBox.warning(
                self,
                "Add Column",
                "Please select a sheet first.",
            )
            return

        column_number = self.table.columnCount() + 1

        self.add_column_button.setEnabled(False)
        self.add_column_button.setText("Adding...")

        try:
            result = self.api_client.add_column(
                file_id=self.file_id,
                sheet_name=sheet_name,
                column_number=column_number,
            )

            worksheet = self.workbook[sheet_name]

            # Backend succeeded, so update local workbook.
            worksheet.insert_cols(
                column_number,
                1,
            )

            # Update table directly. Do not reload the workbook from
            # the server here, because that would cause the visible
            # table to jump/reload.
            self._loading_table = True
            self.table.blockSignals(True)

            try:
                self.table.insertColumn(
                    column_number - 1
                )

                self.table.setHorizontalHeaderItem(
                    column_number - 1,
                    QTableWidgetItem(
                        self.column_letter(column_number)
                    ),
                )

                for row in range(
                    self.table.rowCount()
                ):
                    self.table.setItem(
                        row,
                        column_number - 1,
                        QTableWidgetItem(""),
                    )

            finally:
                self.table.blockSignals(False)
                self._loading_table = False

            self._shift_pending_changes_after_column_insert(
                sheet_name,
                column_number,
            )

            self.table.resizeColumnsToContents()

            QMessageBox.information(
                self,
                "Column Added",
                result.get(
                    "message",
                    "Column added successfully.",
                ),
            )

        except Exception as exc:
            QMessageBox.critical(
                self,
                "Add Column Failed",
                str(exc),
            )

        finally:
            self.add_column_button.setEnabled(True)
            self.add_column_button.setText("Add Column")

    def _shift_pending_changes_after_column_insert(
        self,
        sheet_name: str,
        inserted_column: int,
    ):
        updated = {}

        for (
            change_sheet,
            cell,
        ), value in self.pending_changes.items():

            if change_sheet != sheet_name:
                updated[(change_sheet, cell)] = value
                continue

            letters = ""
            digits = ""

            for char in cell:
                if char.isalpha():
                    letters += char
                elif char.isdigit():
                    digits += char

            if not letters or not digits:
                updated[(change_sheet, cell)] = value
                continue

            old_column = self.column_number(letters)

            if old_column >= inserted_column:
                old_column += 1

            new_cell = (
                self.column_letter(old_column)
                + digits
            )

            updated[
                (change_sheet, new_cell)
            ] = value

        self.pending_changes = updated

    # =========================================================
    # Rename Column
    # =========================================================

    def rename_column(self):
        current_column = self.table.currentColumn()

        if current_column < 0:
            QMessageBox.warning(
                self,
                "Rename Column",
                "Please select a column first.",
            )
            return

        sheet_name = self.selected_sheet()

        if not sheet_name:
            return

        column_number = current_column + 1
        column_letter = self.column_letter(column_number)

        header_item = self.table.horizontalHeaderItem(
            current_column
        )

        current_name = (
            header_item.text()
            if header_item
            else column_letter
        )

        new_name, ok = QInputDialog.getText(
            self,
            "Rename Column",
            (
                f"Enter new name for "
                f"column {column_letter}:"
            ),
            text=current_name,
        )

        if not ok:
            return

        new_name = new_name.strip()

        if not new_name:
            QMessageBox.warning(
                self,
                "Rename Column",
                "Column name cannot be empty.",
            )
            return

        if new_name == current_name:
            QMessageBox.information(
                self,
                "Rename Column",
                "The column name has not changed.",
            )
            return

        self.rename_column_button.setEnabled(False)
        self.rename_column_button.setText("Renaming...")

        try:
            result = self.api_client.update_column(
                file_id=self.file_id,
                sheet_name=sheet_name,
                column_number=column_number,
                column_name=new_name,
            )

            worksheet = self.workbook[sheet_name]

            worksheet.cell(
                row=1,
                column=column_number,
                value=new_name,
            )

            self._loading_table = True
            self.table.blockSignals(True)

            try:
                self.table.setHorizontalHeaderItem(
                    current_column,
                    QTableWidgetItem(new_name),
                )
            finally:
                self.table.blockSignals(False)
                self._loading_table = False

            QMessageBox.information(
                self,
                "Column Renamed",
                result.get(
                    "message",
                    "Column renamed successfully.",
                ),
            )

        except Exception as exc:
            QMessageBox.critical(
                self,
                "Rename Column Failed",
                str(exc),
            )

        finally:
            self.rename_column_button.setEnabled(True)
            self.rename_column_button.setText("Rename Column")

    # =========================================================
    # Delete Column
    # =========================================================

    def delete_column(self):
        current_column = self.table.currentColumn()

        if current_column < 0:
            QMessageBox.warning(
                self,
                "Delete Column",
                "Please select a column first.",
            )
            return

        sheet_name = self.selected_sheet()

        if not sheet_name:
            return

        column_number = current_column + 1
        column_letter = self.column_letter(column_number)

        answer = QMessageBox.question(
            self,
            "Delete Column",
            (
                f"Are you sure you want to delete "
                f"column {column_letter}?"
            ),
            QMessageBox.StandardButton.Yes
            | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )

        if answer != QMessageBox.StandardButton.Yes:
            return

        self.delete_column_button.setEnabled(False)
        self.delete_column_button.setText("Deleting...")

        try:
            result = self.api_client.delete_column(
                file_id=self.file_id,
                sheet_name=sheet_name,
                column_number=column_number,
            )

            worksheet = self.workbook[sheet_name]

            worksheet.delete_cols(
                column_number,
                1,
            )

            self._loading_table = True
            self.table.blockSignals(True)

            try:
                self.table.removeColumn(
                    current_column
                )
            finally:
                self.table.blockSignals(False)
                self._loading_table = False

            self._shift_pending_changes_after_column_delete(
                sheet_name,
                column_number,
            )

            QMessageBox.information(
                self,
                "Column Deleted",
                result.get(
                    "message",
                    "Column deleted successfully.",
                ),
            )

        except Exception as exc:
            QMessageBox.critical(
                self,
                "Delete Column Failed",
                str(exc),
            )

        finally:
            self.delete_column_button.setEnabled(True)
            self.delete_column_button.setText("Delete Column")

    def _shift_pending_changes_after_column_delete(
        self,
        sheet_name: str,
        deleted_column: int,
    ):
        updated = {}

        for (
            change_sheet,
            cell,
        ), value in self.pending_changes.items():

            if change_sheet != sheet_name:
                updated[(change_sheet, cell)] = value
                continue

            letters = ""
            digits = ""

            for char in cell:
                if char.isalpha():
                    letters += char
                elif char.isdigit():
                    digits += char

            if not letters or not digits:
                updated[(change_sheet, cell)] = value
                continue

            old_column = self.column_number(letters)

            if old_column == deleted_column:
                continue

            if old_column > deleted_column:
                old_column -= 1

            new_cell = (
                self.column_letter(old_column)
                + digits
            )

            updated[
                (change_sheet, new_cell)
            ] = value

        self.pending_changes = updated

    @staticmethod
    def column_number(column_letters: str) -> int:
        result = 0

        for char in column_letters.upper():
            if not ("A" <= char <= "Z"):
                continue

            result = (
                result * 26
                + ord(char)
                - ord("A")
                + 1
            )

        return result

    # =========================================================
    # Create Sheet
    # =========================================================

    def create_sheet(self):
        if not self.workbook:
            QMessageBox.warning(
                self,
                "Create Sheet",
                "Workbook is not loaded.",
            )
            return

        sheet_name, ok = QInputDialog.getText(
            self,
            "Create Sheet",
            "Enter new sheet name:",
        )

        if not ok:
            return

        sheet_name = sheet_name.strip()

        if not sheet_name:
            QMessageBox.warning(
                self,
                "Create Sheet",
                "Sheet name cannot be empty.",
            )
            return

        if sheet_name in self.workbook.sheetnames:
            QMessageBox.warning(
                self,
                "Create Sheet",
                (
                    f"Sheet '{sheet_name}' "
                    "already exists."
                ),
            )
            return

        self.create_sheet_button.setEnabled(False)
        self.create_sheet_button.setText("Creating...")

        try:
            result = self.api_client.create_sheet(
                file_id=self.file_id,
                sheet_name=sheet_name,
            )

            self.workbook.create_sheet(sheet_name)

            self.rebuild_sheet_combo(sheet_name)
            self.load_sheet(sheet_name)

            QMessageBox.information(
                self,
                "Sheet Created",
                result.get(
                    "message",
                    "Sheet created successfully.",
                ),
            )

        except Exception as exc:
            QMessageBox.critical(
                self,
                "Create Sheet Failed",
                str(exc),
            )

        finally:
            self.create_sheet_button.setEnabled(True)
            self.create_sheet_button.setText("New Sheet")

    # =========================================================
    # Rename Sheet
    # =========================================================

    def rename_sheet(self):
        if not self.workbook:
            QMessageBox.warning(
                self,
                "Rename Sheet",
                "Workbook is not loaded.",
            )
            return

        current_sheet = self.selected_sheet()

        if not current_sheet:
            QMessageBox.warning(
                self,
                "Rename Sheet",
                "Please select a sheet first.",
            )
            return

        new_sheet_name, ok = QInputDialog.getText(
            self,
            "Rename Sheet",
            "Enter new sheet name:",
            text=current_sheet,
        )

        if not ok:
            return

        new_sheet_name = new_sheet_name.strip()

        if not new_sheet_name:
            QMessageBox.warning(
                self,
                "Rename Sheet",
                "Sheet name cannot be empty.",
            )
            return

        if new_sheet_name == current_sheet:
            QMessageBox.information(
                self,
                "Rename Sheet",
                "The sheet name has not changed.",
            )
            return

        if new_sheet_name in self.workbook.sheetnames:
            QMessageBox.warning(
                self,
                "Rename Sheet",
                (
                    f"Sheet '{new_sheet_name}' "
                    "already exists."
                ),
            )
            return

        self.rename_sheet_button.setEnabled(False)
        self.rename_sheet_button.setText("Renaming...")

        try:
            result = self.api_client.rename_sheet(
                file_id=self.file_id,
                sheet_name=current_sheet,
                new_sheet_name=new_sheet_name,
            )

            worksheet = self.workbook[current_sheet]
            worksheet.title = new_sheet_name

            updated_pending = {}

            for (
                key,
                value,
            ) in self.pending_changes.items():

                old_sheet, cell = key

                if old_sheet == current_sheet:
                    updated_pending[
                        (new_sheet_name, cell)
                    ] = value
                else:
                    updated_pending[key] = value

            self.pending_changes = updated_pending

            self.rebuild_sheet_combo(new_sheet_name)
            self.load_sheet(new_sheet_name)

            QMessageBox.information(
                self,
                "Sheet Renamed",
                result.get(
                    "message",
                    "Sheet renamed successfully.",
                ),
            )

        except Exception as exc:
            QMessageBox.critical(
                self,
                "Rename Sheet Failed",
                str(exc),
            )

        finally:
            self.rename_sheet_button.setEnabled(True)
            self.rename_sheet_button.setText("Rename Sheet")

    # =========================================================
    # Delete Sheet
    # =========================================================

    def delete_sheet(self):
        if not self.workbook:
            QMessageBox.warning(
                self,
                "Delete Sheet",
                "Workbook is not loaded.",
            )
            return

        current_sheet = self.selected_sheet()

        if not current_sheet:
            QMessageBox.warning(
                self,
                "Delete Sheet",
                "Please select a sheet first.",
            )
            return

        if len(self.workbook.sheetnames) == 1:
            QMessageBox.warning(
                self,
                "Delete Sheet",
                (
                    "Cannot delete the only "
                    "sheet in the workbook."
                ),
            )
            return

        answer = QMessageBox.question(
            self,
            "Delete Sheet",
            (
                f"Are you sure you want to "
                f"delete sheet '{current_sheet}'?"
            ),
            QMessageBox.StandardButton.Yes
            | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )

        if answer != QMessageBox.StandardButton.Yes:
            return

        self.delete_sheet_button.setEnabled(False)
        self.delete_sheet_button.setText("Deleting...")

        try:
            result = self.api_client.delete_sheet(
                file_id=self.file_id,
                sheet_name=current_sheet,
            )

            worksheet = self.workbook[current_sheet]
            self.workbook.remove(worksheet)

            self.clear_pending_for_sheet(current_sheet)

            remaining = self.workbook.sheetnames

            self.rebuild_sheet_combo(
                remaining[0] if remaining else None
            )

            if remaining:
                self.load_sheet(
                    self.sheet_combo.currentText()
                )
            else:
                self.table.clear()
                self.table.setRowCount(0)
                self.table.setColumnCount(0)

            QMessageBox.information(
                self,
                "Sheet Deleted",
                result.get(
                    "message",
                    "Sheet deleted successfully.",
                ),
            )

        except Exception as exc:
            QMessageBox.critical(
                self,
                "Delete Sheet Failed",
                str(exc),
            )

        finally:
            self.delete_sheet_button.setEnabled(True)
            self.delete_sheet_button.setText("Delete Sheet")


    # =========================================================
    # Version History
    # =========================================================

    def show_version_history(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Version History")
        dialog.setMinimumSize(850, 500)

        layout = QVBoxLayout()

        title = QLabel(
            f"Version History - {self.filename}"
        )
        title.setStyleSheet(
            """
            QLabel {
                font-size: 20px;
                font-weight: bold;
            }
            """
        )

        layout.addWidget(title)

        versions_table = QTableWidget()
        versions_table.setColumnCount(4)
        versions_table.setHorizontalHeaderLabels(
            [
                "Version",
                "Created At",
                "Created By",
                "Change Summary",
            ]
        )

        versions_table.setEditTriggers(
            QTableWidget.EditTrigger.NoEditTriggers
        )
        versions_table.setSelectionBehavior(
            QTableWidget.SelectionBehavior.SelectRows
        )
        versions_table.setSelectionMode(
            QTableWidget.SelectionMode.SingleSelection
        )

        layout.addWidget(versions_table)

        button_layout = QHBoxLayout()

        restore_button = QPushButton("Restore Selected")
        restore_button.setMinimumHeight(38)

        close_button = QPushButton("Close")
        close_button.setMinimumHeight(38)

        button_layout.addWidget(restore_button)
        button_layout.addStretch()
        button_layout.addWidget(close_button)

        close_button.clicked.connect(dialog.reject)

        layout.addLayout(button_layout)

        dialog.setLayout(layout)

        try:
            result = self.api_client.get_workbook_versions(
                file_id=self.file_id
            )

            versions = result.get("versions", [])

            versions_table.setRowCount(len(versions))

            for row, version in enumerate(versions):
                versions_table.setItem(
                    row,
                    0,
                    QTableWidgetItem(
                        str(version.get("version_number", ""))
                    ),
                )

                versions_table.setItem(
                    row,
                    1,
                    QTableWidgetItem(
                        str(version.get("created_at", ""))
                    ),
                )

                versions_table.setItem(
                    row,
                    2,
                    QTableWidgetItem(
                        str(version.get("created_by", ""))
                    ),
                )

                versions_table.setItem(
                    row,
                    3,
                    QTableWidgetItem(
                        str(version.get("change_summary") or ""),
                    ),
                )

            versions_table.resizeColumnsToContents()

            versions_table.horizontalHeader().setStretchLastSection(True)

            if not versions:
                QMessageBox.information(
                    self,
                    "Version History",
                    "No version history is available for this workbook.",
                )
                return

        except Exception as exc:
            QMessageBox.critical(
                self,
                "Version History",
                f"Failed to load version history:\n\n{exc}",
            )
            return


        def restore_selected_version():
            selected_rows = versions_table.selectionModel().selectedRows()

            if not selected_rows:
                QMessageBox.warning(
                    dialog,
                    "Restore Version",
                    "Please select a version to restore.",
                )
                return

            row = selected_rows[0].row()

            version_item = versions_table.item(row, 0)

            if version_item is None:
                QMessageBox.warning(
                    dialog,
                    "Restore Version",
                    "Unable to determine the selected version.",
                )
                return

            version_number = int(version_item.text())

            reply = QMessageBox.question(
                dialog,
                "Confirm Restore",
                (
                    f"Are you sure you want to restore version "
                    f"{version_number}?\n\n"
                    "Your current workbook state will be replaced "
                    "with this version."
                ),
                QMessageBox.StandardButton.Yes
                | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )

            if reply != QMessageBox.StandardButton.Yes:
                return

            try:
                result = self.api_client.restore_workbook_version(
                    file_id=self.file_id,
                    version_number=version_number,
                )

                QMessageBox.information(
                    dialog,
                    "Restore Successful",
                    (
                        f"Version {version_number} has been restored.\n\n"
                        f"New version: {result.get('new_version', '')}"
                    ),
                )

                dialog.accept()
                self.load_excel_file()

            except Exception as exc:
                QMessageBox.critical(
                    dialog,
                    "Restore Failed",
                    f"Failed to restore version {version_number}:\n\n{exc}",
                )


        restore_button.clicked.connect(restore_selected_version)

        dialog.exec()

    # =========================================================
    # Cleanup
    # =========================================================

    def closeEvent(self, event):
        if self.temp_file_path:
            try:
                os.remove(self.temp_file_path)
            except OSError:
                pass

        event.accept()
